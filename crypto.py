import time
from pycoingecko import CoinGeckoAPI
from openpyxl import Workbook, load_workbook

cg = CoinGeckoAPI()

def fetch_crypto_data():
    """Fetches live cryptocurrency data from CoinGecko API."""
    data = cg.get_coins_markets(vs_currency="usd", order="market_cap_desc", per_page=50, page=1)
    return [{
        "name": crypto["name"],
        "symbol": crypto["symbol"].upper(),
        "price": crypto["current_price"],
        "market_cap": crypto["market_cap"],
        "volume": crypto["total_volume"],
        "price_change": crypto["price_change_percentage_24h"]
    } for crypto in data]

def update_excel():
    """Fetches live data, updates the Excel file, and performs basic analysis."""
    crypto_data = fetch_crypto_data()

    try:
        workbook = load_workbook("Crypto.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Cryptocurrency Name", "Symbol", "Current Price (USD)", "Market Capitalization", "24h Trading Volume", "Price Change (24h %)"])

    # Clear old data before adding new data
    sheet.delete_rows(2, sheet.max_row)

    # Add new live data
    for crypto in crypto_data:
        sheet.append([crypto["name"], crypto["symbol"], crypto["price"], crypto["market_cap"], crypto["volume"], crypto["price_change"]])

    # Perform analysis
    top_5 = sorted(crypto_data, key=lambda x: x["market_cap"], reverse=True)[:5]
    avg_price = sum(crypto["price"] for crypto in crypto_data) / len(crypto_data)
    highest_change = max(crypto_data, key=lambda x: x["price_change"])
    lowest_change = min(crypto_data, key=lambda x: x["price_change"])

    # Append analysis results
    sheet.append([])
    sheet.append(["Top 5 Cryptos by Market Cap"])
    for crypto in top_5:
        sheet.append([crypto["name"], crypto["market_cap"]])

    sheet.append(["Average Price of Top 50", avg_price])
    sheet.append(["Highest 24h Change", highest_change["name"], highest_change["price_change"]])
    sheet.append(["Lowest 24h Change", lowest_change["name"], lowest_change["price_change"]])

    workbook.save("Crypto.xlsx")
    print("Excel file updated successfully!")

# Run the script continuously every 5 minutes
while True:
    update_excel()
    print("Waiting for the next update...\n")
    time.sleep(300)  # Update every 5 minutes
